from qiskit import execute, transpile
from qiskit.circuit import QuantumCircuit
from qiskit.quantum_info import Operator # Used to define quantum gates via matrices
from qiskit_aer import AerSimulator, Aer
from qiskit.tools.parallel import parallel_map
import matplotlib.pyplot as plt
from qiskit.visualization import plot_histogram
import numpy as np
import warnings # Used to suppress warning of "execute" being deprecated
import sys # Used to access cmd-line args
import time # Used to measure timecost of functions
import scipy.stats as scipy
from openpyxl import Workbook
from openpyxl import load_workbook
from multiprocessing import Pool
from threading import Lock
from enum import Enum

s_print_lock = Lock()

ShotsDefinition = 10000
RepeatsOfSameExperiment = 100
DATAPATH = "data/newExperiments/"

class Iterations(Enum):
    MAXIMUM = 1
    SUBOPTIMAL = 2

    def __str__(self):
        if self == Iterations.MAXIMUM: return "MAXIMUM"
        if self == Iterations.SUBOPTIMAL: return "SUBOPTIMAL"



# Generate {k} marked items in the range [0,{n}]
def getMarkedItems(n, k):
    elements = list(range(n))
    marked = []
    for _ in range(k):
        index = np.random.randint(len(elements))
        marked.append(elements[index])
        del elements[index]
    return marked

# Create an Oracle circuit for {numberOfQubits} and the indices in the array {marked}
def oracleFactory(numberOfQubits, marked, gateName="Delphi"):
    oracle = QuantumCircuit(numberOfQubits, name=gateName)
    oracle_matrix = np.identity(2**numberOfQubits)
    for index_to_mark in marked:
        oracle_matrix[index_to_mark, index_to_mark] = -1
    oracle.unitary(Operator(oracle_matrix), range(numberOfQubits))
    #print(oracle.decompose().draw())
    return oracle.to_gate()

# Create a Diffusion circuit for {numberOfQubits}
def diffusionFactory(numberOfQubits, gateName="Diffusion"):
    phaseOr = QuantumCircuit(numberOfQubits, name=gateName)
    phaseOr_matrix = -1*np.identity(2**numberOfQubits) # Phase flip everthing except the |0> state
    phaseOr_matrix[0,0] = 1
    phaseOr.unitary(Operator(phaseOr_matrix), range(numberOfQubits))
    #print(phaseOr.decompose().draw())
    return phaseOr.to_gate()

# Create a Grover-operation circuit for {numberOfQubits} and the {marked} items
def groverOperation(numberOfQubits, marked, gateName="Grover Op({},{})"):
    oracle = oracleFactory(numberOfQubits, marked, "Delphi({}, {})".format(numberOfQubits, len(marked)))
    phaseOr = diffusionFactory(numberOfQubits, "Diffusion({}, {})".format(numberOfQubits, len(marked)))

    groverOp = QuantumCircuit(numberOfQubits, name=gateName.format(numberOfQubits, len(marked)))
    groverOp.append(oracle, range(numberOfQubits))
    groverOp.h(range(numberOfQubits))
    groverOp.append(phaseOr, range(numberOfQubits))
    groverOp.h(range(numberOfQubits))
    #print(groverOp.decompose().draw())
    return groverOp.to_gate()

# Determine the optimal number of iterations to perform for the instance ({n},{k})
def groverIterations(n, k):
    return int(np.pi/(4 * np.arcsin(np.sqrt(np.divide(k,n)))))

# Determine the suboptimal number of iterations to perform for the instance ({n},{k}), which may give fewer iterations in the long run
def groverSuboptimalIterations(n,k):
    return int(np.round(0.58278*np.sqrt(n/k)))

# Generate a statevector for the given {completeCircuit}
def checkState(completeCircuit):
    backend = Aer.get_backend('statevector_simulator')
    with warnings.catch_warnings(action="ignore"):
        job = execute(completeCircuit, backend)

    result = job.result()
    sv = result.get_statevector()
    print(np.around(sv,2))
    return sv


def runState(qubits, k, shots, repetitions):
    if(qubits > 10):
        print("You are requesting too many qubits! Maximum of 10")
        return
    n = 2**qubits
    groverIts = groverIterations(n,k)
    markedItems = getMarkedItems(n,k)
    print("(N,k)=({},{}), m={}, Marked items: ".format(n,k,groverIts), markedItems)

    groverOp = groverOperation(qubits, markedItems)
    grover = QuantumCircuit(qubits,qubits)
    grover.h(range(qubits)) # Initial state
    for _ in range(groverIts):
        grover.append(groverOp, range(qubits))

    sv=checkState(grover)
    print(grover_success_prob(n, k, groverIts), (sv[markedItems[0]].real**2)*k)


# Run the given {completeCircuit} with {numberOfQubits} for {numberOfShots}
def runCircuit(completeCircuit, numberOfQubits, numberOfShots):
    backend = Aer.get_backend("qasm_simulator")
    completeCircuit.measure(range(numberOfQubits), range(numberOfQubits))

    with warnings.catch_warnings(action="ignore"): # Ignore the warning of "execute" being deprecated
        job = execute(completeCircuit, backend, shots=numberOfShots)

    result = {int(k,2) : v for k,v in job.result().get_counts().items()}
    return result
    print("Results: ", result)
    plotResults(result)

# Deprecated/useless, but has important comments
def runCircuitManually(completeCircuit, numberOfQubits):
    # This is really slow
    # Possible optimisation:
    #   When we calculate the queries necessary for a successful (N,k) instance
    #   we could 1. Perform multiple shots, 2. Divide m (number of iterations)*numberOfShots by the number of successful shots
    #   This would simulate the necessary "extra" iterations while allowing us to use the "shots" functionality
    #
    #   Restated: The average number of iterations is given by (1+#fails/#successes)*m, regardless of grouping into experiments or not
    #   The problem is we may not analyze medians, modes, or deviations this way
    #       Variation may be approximated via (sannstat formelblad kap 8)


    backend = Aer.get_backend("qasm_simulator")
    completeCircuit.measure(range(numberOfQubits), range(numberOfQubits))
    #print(completeCircuit.draw())

    measurements = {}

    for _ in range(numberOfShots):
        with warnings.catch_warnings(action="ignore"):
            job = execute(completeCircuit, backend, shots=1)
        result = job.result()
        for k,d in result.get_counts().items():
            if k in measurements.keys():
                measurements[k] += d
            else:
                measurements[k] = d

    print("Results: ", measurements)
    plotResults(measurements)

# Plot the {plotData} of an execution of an ({n},_) instance with the {marked} elements. Red bars are the marked ones, the blue are the unmarked ones
def plotResults(plotData, n, marked):
    cols = ["r" if x in marked else "b" for x in plotData.keys()]
    plt.bar(plotData.keys(), plotData.values(), color=cols)
    plt.show()
    
    

# Main experiment function: create and run a (2**{qubits}, {k}) instance for {shots}, and repeat {repetitions} times to calculate averages
def runGroverExperiment(qubits, k, iterations, shots, repetitions):
    if(qubits > 10):
        print("You are requesting too many qubits! Maximum of 10")
        return 0,0
    n = 2**qubits
    groverIts = iterations(n,k)
    markedItems = getMarkedItems(n,k)

    groverOp = groverOperation(qubits, markedItems)
    grover = QuantumCircuit(qubits,qubits)
    grover.h(range(qubits)) # Initial state
    for _ in range(groverIts):
        grover.append(groverOp, range(qubits))


    correctMeasurements = np.empty(repetitions) # This will be Binomially distributed with Bin(n, grov_succ_prob(groverIts, k, n))
    averageQueryCounts = np.empty(repetitions) # averageQueryCounts[i] := average number of Oracle calls needed when measuring one of correctMeasurements[i]
    
    for i in range(repetitions):
        result = runCircuit(grover, qubits, shots)

        correctMeasurements[i] = sum(result[x] for x in list(markedItems & result.keys())) # We can see it as if we performed this many successful measurements
        averageQueryCounts[i] = shots*(groverIts+1)/correctMeasurements[i] # This is then the average number of queries needed for each successful measurements, considering failed measurements as being added onto the count

    


    corr = correctMeasurements.tolist()
    corr.insert(0, "({},{})".format(n,k)) # Append a label so the first column specifies which (N,k) instance the row relates to
    corr.insert(0, "") # Append a label so the first column specifies which (N,k) instance the row relates to
    avg = averageQueryCounts.tolist()
    avg.insert(0, "({},{})".format(n,k)) # Append a label so the first column specifies which (N,k) instance the row relates to
    avg.insert(0, "") # Append a label so the first column specifies which (N,k) instance the row relates to

    return corr,avg

# Calculate the mathematical probability of correct measurement for an ({n},{k}) instance after {r} Grover iterations
def grover_success_prob(n,k,r):
    return np.sin((2*r + 1) * np.arcsin(np.sqrt(k/n)))**2#*100          


def runSimulationN(n, wb, fn, suffix, iterations):
    # Create the sheet where the data from this test should be written
    ws1 = wb.create_sheet("N={},Correct({})".format(n,suffix))
    ws2 = wb.create_sheet("N={},Avg({})".format(n,suffix))


    ws1.append([""])
    ws2.append([""])


    # Run each (N,k) test for the specified N
    for k in range(1,int((2**n)/2)+1):
        correctLine,avgLine = runGroverExperiment(n,k,iterations,ShotsDefinition,RepeatsOfSameExperiment)    # Get the data from the (N,k) tests
        ws1.append(correctLine)     # Append the data as a new line to the worksheet
        ws2.append(avgLine)

        if k % 50 == 0:
            wb.save(fn)


# Takes a list of numbers of qubits and performs tests with it.
# Each qubit gets its own worksheet in a shared workbook
def parallelSimulate(args):
    # Create the filename. Currently for a list [q1, q2, q3] the name becomes "q1_q2_q3.xlsx"
    fileName = DATAPATH + '_'.join(str(x) for x in args) + ".xlsx"
    wb = Workbook()
    wb.remove(wb['Sheet']) # Remove the default worksheet
    # For each qubit, run the test
    s_print(">Thread for {} has started executing".format(fileName))

    if args[0] == Iterations.MAXIMUM:
        for q in args[1:]:
            runSimulationN(q, wb,fileName, "MAXIMUM", groverIterations)
            wb.save(fileName)
            s_print("Simulations for {} qubits (MAXIMUM) are done".format(q))
    elif args[0] == Iterations.SUBOPTIMAL:
        for q in args[1:]:
            runSimulationN(q, wb,fileName, "SUBOPTIMAL", groverSuboptimalIterations)
            wb.save(fileName)
            s_print("Simulations for {} qubits (SUBOPTIMAL) are done".format(q))
    else:
        s_print("Error first argument: ", args)


    s_print(">Thread for {} has finished executing".format(fileName))

    return 1


    
def s_print(*a, **b):
    """Thread safe print function"""
    with s_print_lock:
        print(*a, **b)


def main():
    # Each element is the argument to a single thread. If one thread should simulate 1,3, and 6 qubits its list should be [1,3,6]
    threadArgs = [ 
                   [Iterations.MAXIMUM,10]]
    with Pool() as pool:
      result = pool.map(parallelSimulate, threadArgs)
    print("Program finished!")

    #runState(qubits=4, k=5, shots=ShotsDefinition, repetitions=RepeatsOfSameExperiment)
    return







if __name__ == "__main__":
    main()

    # estProb = meanOfCorrectMeasurements/repetitions # Estimated successprobability of a measurement, can be compared with actual successprobability
    # estQueryVariance = sum(x**2 for x in (averageQueryCounts - meanOfAverageQueries))/(repetitions-1) # Variance of average queries per successful measurement
    # estSuccessVariance = sum(x**2 for x in (correctMeasurements - meanOfCorrectMeasurements))/(repetitions-1)

    

    # #print("Successrate: {}/{}, avg oracles={}, oracleVariance={}, successVar={}".format(meanOfCorrectMeasurements, shots, meanOfAverageQueries, estQueryVariance, estSuccessVariance))
    # mean, var = scipy.binom.stats(shots, grover_success_prob(n, k, groverIts))
    # #print("Mean: {}, Var: {}".format(mean, var))
    # #print("Results: ", result)
    # #plotResults(result, n, markedItems)

    # #data = interweave(correctMeasurements,averageQueryCounts).tolist()

# def test():
#     minimumIter = 0
#     for q in range(2,11):
#         n = 2**q
#         for k in range(1, int(n/2)):
#             m1 = groverIterations(n,k) 
#             m2= groverSuboptimalIterations(n,k)
#             if np.abs(m1-m2) > 0:
#                 print("({},{}): {}, {}".format(n,k,m1,m2))

