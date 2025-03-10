import numpy as np
import math
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import statistics
# Determine the optimal number of iterations to perform for the instance ({n},{k})
def groverIterations(n, k):
    return int(np.pi/(4 * np.arcsin(np.sqrt(np.divide(k,n)))))

# Determine the suboptimal number of iterations to perform for the instance ({n},{k}), which may give fewer iterations in the long run
def groverSuboptimalIterations(n,k):
    return int(np.round(0.58278*np.sqrt(n/k)))

def getM(n,k):
    for m in range(k,n):
        if (math.factorial(m)*math.factorial(n-k))/(math.factorial(m-k)*math.factorial(n)) >= 0.5:
            print("({},{}) {}, {}".format(n,k, (math.factorial(m)*math.factorial(n-k))/(math.factorial(m-k)*math.factorial(n)), (math.factorial(m-1)*math.factorial(n-k))/(math.factorial(m-k-1)*math.factorial(n))))
            return m

def fix(correct,avg,target):
    shots=10000
    target.append([""])
    iter = 0
    skip=True
    for i,j in zip(correct.rows, avg.rows):
        if skip:
            skip = False
            continue
        corrRow = list(x.value for x in i)
        avgRow = list(x.value for x in j)
        tRow = list(avg + (shots/x) for x,avg in zip(corrRow[2:],avgRow[2:]))
        tRow.insert(0, avgRow[1])
        tRow.insert(0, "")
        target.append(tRow)

def variance():
    outWb = Workbook()
    outWb.remove(outWb['Sheet'])
    
    for fn in ["newExperiments/MAXIMUM_10.xlsx", "modified/SUBOPTIMAL_10.xlsx"]:
        wb = load_workbook("data/" + fn, data_only=True)
        
        for sheet in wb.worksheets:
            # Hitta sheetet med Avg värdena (de vi använder)
            if (sheet.title.__contains__("Corrected")):
                outsheet = outWb.create_sheet("Avg" + fn.split('/')[0])
                skip = True
                sums = np.zeros(100) # Här kommer kolmnerna att summeras var för sig
                for row in sheet:
                    if skip:
                        skip = False
                        continue
                    avg = row[0].value # Rad-medel finns redan så jag använder det
                    label = row[1].value
                    var = sum([(a.value - avg)**2 for a in row[2:]])/(100-1) # Stickprovsvariansen. Är detta rätt? Tror det, men sanity checka mig snälla
                    sums += np.array([a.value for a in row[2:]]) # Summera kolumnerna var för sig

                    outsheet.append([label, var])
                outline = list(sums)
                outline.insert(0, "Total extraction instances")
                outsheet.append(outline)

                avg = statistics.mean(sums)
                var = sum([(a-avg)**2 for a in sums])/(100-1) # Stickprovsvarians av kolumnsummorna (aka alla enskilda totala extraktioner)
                outsheet.append(["Variance of total extraction", var])

            elif (sheet.title.__contains__("Correct(")): # Hitta sheetet med orignalvärdena (de rakt från Qiskit)
                skip = True
                outsheet = outWb.create_sheet("Measures" + fn.split('/')[0])
                for row in sheet:
                    if skip:
                        skip = False
                        continue

                    label = row[1].value
                    data = [a.value for a in row[2:]] # Skaffa radvärdena
                    var = sum([(a - statistics.mean(data))**2 for a in data])/(100-1) # Stickprovsvarians av Qiskit resultaten
                    outsheet.append([label, var])

            
            
    outWb.save("data/variance.xlsx")
def main():
    for q in range(1,11):
        n = int(2**q)
        for k in range(1,int(n/2)+1):
            m1 = groverIterations(n,k)
            m2 = groverSuboptimalIterations(n,k)
            if m1-m2 > 2:
                print("({},{}): Opt: {}, Pess: {}".format(n,k,m1,m2))
    return()
            

            

    return
    fileName = "data/averageClassical.xlsx"
    wb = Workbook()
    wb.remove(wb['Sheet']) # Remove the default worksheet
    ws1 = wb.create_sheet("OracleCalls")
    ws1.append([""])

    for q in range(10,11):
        n = int(2**q)
        for k in range(1,int(n/2)+1):
            m = getM(n,k)
            ws1.append(["({},{})".format(n,k), m])
    #wb.save(fileName)


if __name__ == "__main__":
    main()