from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
from qiskit.tools.parallel import parallel_map

reps = 100

# Takes two lists and interweaves them, so the related datapoints end up in neighbouring columns in the Excel sheet
# Stolen from: https://stackoverflow.com/questions/5347065/interleaving-two-numpy-arrays-efficiently/5347492#5347492
def interweave(a,b):
    c = np.empty((a.size + b.size,), dtype=a.dtype)
    c[0::2] = a
    c[1::2] = b
    return c

# Test-function that "emulates" the {runGroverExperiment} function
def test(n,k):
    
    correctMeasurements = np.array(range(reps)) 
    averageQueryCounts = np.array(range(reps)) + 100*k
    
    data = interweave(correctMeasurements,averageQueryCounts).tolist()

    data.insert(0, "({},{})".format(n,k)) # Append a label so the first column specifies which (N,k) instance the row relates to

    return data

def runTest(n, wb):
    # Create the sheet where the data from this test should be written
    ws = wb.create_sheet("N={}".format(n))

    # Generate header line for the worksheet
    firstLine = [""] # First cell blank? 
    for i in range(reps):
        firstLine.append("correctMeasure{}".format(i))
        firstLine.append("averageQueryCount{}".format(i))

    ws.append(firstLine)

    # Run each (N,k) test for the specified N
    for k in range(1,int(n/2)+1):
        line = test(n,k)    # Get the data from the (N,k) tests
        ws.append(line)     # Append the data as a new line to the worksheet
    
# Takes a list of numbers of qubits and performs tests with it.
# Each qubit gets its own worksheet in a shared workbook
def parallelTest(qubits):
    # Create the filename. Currently for a list [q1, q2, q3] the name becomes "q1_q2_q3.xlsx"
    fileName = '_'.join(str(x) for x in qubits) + ".xlsx"
    wb = Workbook()

    # For each qubit, run the test
    for q in qubits:
        runTest(2**q, wb)
        wb.save("data/" + fileName)

    wb.remove(wb['Sheet']) # Remove the default worksheet
    


def main():
    # These can be performed in parallel
    threadArgs = [[1,2,3], [4,5], [6]]
    parallel_map(parallelTest, threadArgs)

    
    return







if __name__ == "__main__":
    main()